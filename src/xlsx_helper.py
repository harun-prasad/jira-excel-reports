from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle

import datetime
import utils.general_utils
import jx_config
import math
import task_helper

task_helper_obj = task_helper.TaskHelper.get_instance()

thin_side = Side(border_style="thin", color="000000")

datestyle = NamedStyle(name="datestyle")
datestyle.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
datestyle.fill = PatternFill("solid", fgColor="fcd5b4")
datestyle.alignment = Alignment(horizontal= "center", vertical= "center", wrap_text=True)

userstyle = NamedStyle(name="userstyle")
userstyle.fill = PatternFill("solid", fgColor="c5d9f1")
userstyle.alignment = Alignment(vertical= "center", wrap_text=True)
userstyle.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

taskstyle = NamedStyle(name="taskstyle")
taskstyle.fill = PatternFill("solid", fgColor="99ffcc")
taskstyle.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
taskstyle.alignment = Alignment(horizontal= "center", vertical= "center", wrap_text=True)

notaskstyle = NamedStyle(name="notaskstyle")
notaskstyle.fill = PatternFill("solid", fgColor="e6b8b7")
notaskstyle.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

taskdetailstyle = NamedStyle(name="taskdetailstyle")
taskdetailstyle.alignment = Alignment(vertical= "center", wrap_text=True)
taskdetailstyle.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

class XlsxHelper:
    def __init__(self, name="jira_worklog.xlsx"):
        """XLSX Helper initialisation function
        
        Keyword Arguments:
            name {str} -- Output Excel file name (default: {"jira_worklog.xlsx"})
        """                
        self.name = name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.user_list = []
        self.work_log_start_date = None
        self.work_log_end_date = None
        self.worklogged_dates = []

    def create_report_template(self, user_list, work_log_start_date, work_log_end_date, worklogged_dates):
        """[summary]
        Arguments:
            user_list {[type]} -- [description]
            work_log_start_date {[type]} -- [description]
            work_log_end_date {[type]} -- [description]
            worklogged_dates {[type]} -- [description]
        """
        self.user_list = user_list
        self.work_log_start_date = work_log_start_date
        self.work_log_end_date = work_log_end_date
        self.worklogged_dates = worklogged_dates
        # Setting user column
        ws = self.ws
        cell = ws.cell(row= 1, column = 1)
        cell.value = "User"
        cell.style = userstyle
        ws.column_dimensions["A"].width = 20
        start_row = 2
        start_column = 1
        for user in user_list:
            cell = ws.cell(row=start_row, column = start_column)
            cell.value = user
            cell.style = userstyle
            ws.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row + 1, end_column=start_column)
            start_row += 2
        # Setting date row
        start_row = 1
        start_column = 2
        for worklog_date in utils.general_utils.daterange(work_log_start_date,work_log_end_date):
            cell = ws.cell(row=start_row, column=start_column)
            cell.value= datetime.datetime.strftime(worklog_date, "%A, %Y-%m-%d")
            cell.style = datestyle
            if datetime.datetime.strftime(worklog_date,"%Y-%m-%d") in worklogged_dates:
                ws.merge_cells(start_row = start_row, start_column = start_column, end_row=start_row, end_column = start_column + 7)
                start_column += 8
            else:
                ws.merge_cells(start_row = start_row, start_column = start_column, end_row=start_row + len(user_list)*2, end_column = start_column)
                start_column += 1

    def populate_report(self, userwise_worklog, worklogged_dates):
        ws = self.ws
        start_row = 0
        for user in self.user_list:
            print(user)
            start_row += 2
            start_column = 2
            worklog_text_column = 2
            text_lines = 1  # To store the current maximum number of lines in the jira text cell
            #if user not in userwise_worklog:
            #    continue
            # Initializing the current_user task variables used for merging/joinig the task cells in previous day and next day
            user_current_task = ""
            user_next_task = ""
            user_next_task_hours = 0
            # Initializing excel write row and column positions
            for worklog_date in utils.general_utils.daterange(self.work_log_start_date,self.work_log_end_date):
                print(worklog_date)
                day_worklog_hours_filled = 0
                worklog_date_str = datetime.datetime.strftime(worklog_date,"%Y-%m-%d")
                if worklog_date_str not in worklogged_dates:
                    worklog_text_column += 1
                    start_column = worklog_text_column
                    continue
                elif user not in userwise_worklog or worklog_date_str not in userwise_worklog[user]:
                    # TODO No work log Mark in light Red
                    # user_current_task = "" # resetting current and next task merge to nothing
                    # user_next_task = "" # TODO might have to check if this need to be reset
                    cell = ws.cell(row=start_row, column=start_column)
                    cell.style = notaskstyle
                    ws.merge_cells(start_row = start_row, start_column = start_column, end_row = start_row, end_column = start_column + 7)
                    ws.merge_cells(start_row = start_row+1, start_column = start_column, end_row = start_row+1, end_column = start_column + 7)
                    worklog_text_column +=8
                    start_column = worklog_text_column
                    continue
                #work log present
                sorted_tasks = task_helper_obj.sort_task_by_work_log_hours(userwise_worklog[user][worklog_date_str])
                day_total_worklog_hours = task_helper_obj.get_day_total_worklog_hours(userwise_worklog[user][worklog_date_str])
                day_worklog_hours_to_fill = min(8,math.ceil(day_total_worklog_hours))
                # Logic for user_next_task
                i = 1
                while True:
                    next_worklog_date = worklog_date + datetime.timedelta(i)
                    next_worklog_date_str = datetime.datetime.strftime(next_worklog_date, "%Y-%m-%d")
                    if next_worklog_date_str in userwise_worklog[user] or next_worklog_date > self.work_log_end_date :
                        break
                    i += 1
                if user_current_task:
                    start_row, start_column, day_worklog_hours_filled, text_lines = self.fill_task_worklog(user_current_task, userwise_worklog[user][worklog_date_str][user_current_task], start_row, start_column, worklog_text_column, day_worklog_hours_filled, day_worklog_hours_to_fill, text_lines)

                if next_worklog_date_str in userwise_worklog[user]:
                    user_next_task = task_helper_obj.find_task_across_days(userwise_worklog[user][worklog_date_str], userwise_worklog[user][next_worklog_date_str])
                    if user_next_task and user_next_task != user_current_task:
                        user_next_task_hours = min(day_worklog_hours_to_fill,math.ceil(userwise_worklog[user][worklog_date_str][user_next_task]["worklog_hours"]))
                        day_worklog_hours_filled += user_next_task_hours
                # End if
                
                for task in sorted_tasks:
                    task_id = task[0]
                    #if day_worklog_hours_filled == day_worklog_hours_to_fill:
                    #    break
                    if task_id == user_next_task or task_id == user_current_task:
                        continue
                    start_row, start_column, day_worklog_hours_filled, text_lines = self.fill_task_worklog(task_id, userwise_worklog[user][worklog_date_str][task_id], start_row, start_column, worklog_text_column, day_worklog_hours_filled, day_worklog_hours_to_fill, text_lines)
                if user_next_task and user_current_task != user_next_task:
                    day_worklog_hours_filled -= user_next_task_hours
                    start_row, start_column, day_worklog_hours_filled, text_lines = self.fill_task_worklog(user_next_task, userwise_worklog[user][worklog_date_str][user_next_task], start_row, start_column, worklog_text_column, day_worklog_hours_filled, day_worklog_hours_to_fill, text_lines)

                user_current_task = user_next_task
                user_next_task = ""
                if start_column < worklog_text_column+8:
                    cell = ws.cell(row = start_row, column = start_column)
                    cell.style = notaskstyle
                    ws.row_dimensions[start_row].height = 14.4 * 2
                    ws.merge_cells(start_row = start_row, start_column = start_column, end_row = start_row, end_column = worklog_text_column + 7)
                worklog_text_column += 8
                start_column = worklog_text_column
            # End For

    def fill_task_worklog(self, task_id, worklog, start_row, start_column, worklog_text_column, day_worklog_hours_filled, day_worklog_hours_to_fill, text_lines):
        print(task_id)
        ws = self.ws
        task = task_helper_obj.get_task(task_id)

        worklog_text_str_cell = ws.cell(row = start_row + 1, column = worklog_text_column)
        worklog_text_str_cell.style = taskdetailstyle
        if worklog_text_str_cell.value:
            worklog_text_str = worklog_text_str_cell.value[:-4] + "\n" # Removing the text "\n===" added to the bottom or text line.
        else:
            worklog_text_str = ""
        worklog_text_str_cell.value = worklog_text_str + task_id +" (" + task.task_type + ") " + task.name+ "\n" + worklog["worklog_text"] + "\n==="

        no_of_lines = worklog_text_str_cell.value.count('\n') + 1
        if no_of_lines > text_lines:
            text_lines = no_of_lines
            if text_lines > 10: text_lines = 10 # Maxing to 10 lines of height
            ws.row_dimensions[start_row+1].height = 14.4 * text_lines # Value 14.4. taken from Excel for default height of cell for default font

        ws.merge_cells(start_row = start_row+1, start_column=worklog_text_column, end_row=start_row+1, end_column= worklog_text_column + 7)

        if day_worklog_hours_filled < day_worklog_hours_to_fill:
            task_hours = worklog["worklog_hours"]
            if  task_hours > 2:
                task_str = task_id + " "+ task.name
            else:
                task_str = task_id
            fill_length = min(math.ceil(task_hours), day_worklog_hours_to_fill-day_worklog_hours_filled)
            cell = ws.cell(row = start_row, column = start_column)
            cell.value = task_str
            cell.style = taskstyle
            ws.row_dimensions[start_row].height = 14.4 * 2
            
            ws.merge_cells(start_row = start_row, start_column=start_column, end_row=start_row, end_column=start_column + fill_length - 1)
            start_column += fill_length
            day_worklog_hours_filled += fill_length

        return start_row, start_column, day_worklog_hours_filled, text_lines
    
    def save(self):
        self.wb.save(self.name)
    
    def __del__(self):
        #self.save()
        pass

if __name__ == "__main__":
    objxlsx = XlsxHelper()
    objxlsx.create_report_template(jx_config.jira_worklog_users,jx_config.work_log_start_date, jx_config.work_log_end_date, ["2019-12-11"])